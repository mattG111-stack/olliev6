"""The whole run, as a workbook.

Five sheets, because there are five different questions and answering them on
one sheet answers none of them well:

  What happened      the log, in order. Read top to bottom and the run tells
                     its own story: loaded, rejected these for that reason,
                     enriched, blocked here, priced, published.
  Why rows dropped   the count at every gate between "a listing" and "a deal",
                     the same funnel the publish page draws.
  Every listing      one row per property with the decisions attached — what it
                     was valued at, which path produced that number, how many
                     comparable sales stood behind it, why it is held, why it
                     carries no deal signal.
  The deals          just the listings that made it, biggest margin first.
  Held back          just the ones that didn't, with the reason.

Dates are written as real dates in day/month/year, not as text and never as the
ISO string — a spreadsheet that cannot sort its own date column is a screenshot
with extra steps.
"""
from __future__ import annotations

import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .deal_funnel import deal_funnel
from .models import ImportBatch, PropertyForSale
from .runlog import events

# Scraped text carries control characters — a stray \x00 or \x07 inside an
# address, a hold reason, a property type — and openpyxl refuses to write one at
# all: IllegalCharacterError, which surfaced as "Download failed (500)" with
# nothing saying why. ONE bad byte in one cell lost the entire workbook.
#
# The character is worthless; the row is not. Strip and carry on.
_ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
# Excel's own hard limit on a cell. A description field can exceed it.
_MAX_CELL = 32_000


def _clean(v):
    if isinstance(v, str):
        v = _ILLEGAL.sub("", v)
        if len(v) > _MAX_CELL:
            v = v[:_MAX_CELL] + "…"
    return v


_HEAD_FILL = PatternFill("solid", fgColor="2F5D50")
_HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
_MONEY = "#,##0"
_PCT = "0.0%"
_WHEN = "d mmm yyyy h:mm"


def _append(ws, values) -> None:
    """The one door every row goes through, so nothing reaches a cell unchecked."""
    ws.append([_clean(v) for v in values])


def _sheet(wb: Workbook, title: str, headers: list[str], first: bool = False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title[:31]
    _append(ws, headers)
    for c in ws[1]:
        c.fill, c.font = _HEAD_FILL, _HEAD_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    return ws


def _fit(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt(ws, col: int, code: str, start: int = 2) -> None:
    for row in ws.iter_rows(min_row=start, min_col=col, max_col=col):
        row[0].number_format = code


def build(db: Session, batch_id: int) -> bytes:
    b = db.get(ImportBatch, batch_id)
    wb = Workbook()

    # ---- 1. What happened -------------------------------------------------
    ws = _sheet(wb, "What happened",
                ["When", "Stage", "", "What happened", "How many", "Listing"],
                first=True)
    for e in events(db, batch_id):
        _append(ws, [e.at, (e.stage or "").title(),
                   {"warn": "check", "error": "FAILED"}.get(e.level, ""),
                   e.detail or e.event, e.count, e.address])
    _fmt(ws, 1, _WHEN)
    _fmt(ws, 5, _MONEY)
    _fit(ws, [19, 10, 8, 88, 11, 40])

    # ---- 2. Why rows dropped ---------------------------------------------
    f = deal_funnel(db, batch_id)
    ws = _sheet(wb, "Why rows dropped",
                ["Step", "Listings left", "Dropped here", "Why this step drops rows"])
    _append(ws, ["Listings in this load", f.total, None, ""])
    for s in f.steps:
        _append(ws, [s.label, s.kept, s.lost or None, s.why])
    if f.mismatch:
        _append(ws, [])
        _append(ws, ["PASSES EVERY TEST ABOVE AND IS STILL NOT FLAGGED",
                   f.mismatch, None,
                   "The figures on these rows say the deal is there and the flag "
                   "beside them says it does not. Re-run pricing on this load."])
        for x in f.mismatch_examples:
            _append(ws, ["", None, None, x])
    if f.hold_reasons:
        _append(ws, [])
        _append(ws, ["HELD BACK, BY REASON", None, None, ""])
        for reason, n in f.hold_reasons:
            _append(ws, [reason, None, n, ""])
    _fmt(ws, 2, _MONEY)
    _fmt(ws, 3, _MONEY)
    _fit(ws, [46, 14, 14, 92])

    # ---- 3, 4, 5. The listings themselves ---------------------------------
    cols = ["Address", "Suburb", "Type", "Asking", "Where the asking came from",
            "Was listed at", "On", "Council valuation",
            "Our valuation", "Margin $", "Margin %", "Buy price", "Comps",
            "Confidence", "How it was valued", "Deal?", "Why no deal",
            "Held?", "Why held", "Floor m2", "Land m2", "Looked up on"]
    widths = [40, 18, 16, 13, 24, 14, 13, 15, 14, 13, 10, 13, 8, 11, 20, 8, 62,
              8, 34, 9, 9, 15]

    def row(p: PropertyForSale) -> list:
        margin_d = ((p.fair_value - p.asking_price)
                    if (p.fair_value and p.asking_price) else None)
        return [p.address, p.suburb, p.property_type, p.asking_price,
                p.asking_basis, p.prior_asking_price, p.prior_asking_seen_at,
                p.cv_numeric,
                p.fair_value, margin_d, p.margin, p.buy_price, p.comps_used,
                (p.confidence or "").title(), p.expected_sale_path,
                "YES" if p.is_underpriced else "", p.deal_block_reason,
                "held" if p.is_held else "", p.hold_reason,
                p.floor_area_m2, p.land_area_m2, p.pv_checked_at]

    rows = list(db.query(PropertyForSale)
                .filter(PropertyForSale.import_batch_id == batch_id)
                .order_by(PropertyForSale.address))

    for title, subset, order in [
        ("Every listing", rows, None),
        ("The deals", [p for p in rows if p.is_underpriced],
         lambda p: -(p.margin or 0)),
        ("Held back", [p for p in rows if p.is_held], lambda p: p.hold_reason or ""),
    ]:
        ws = _sheet(wb, title, cols)
        for p in (sorted(subset, key=order) if order else subset):
            _append(ws, row(p))
        for c in (4, 6, 8, 9, 10, 12):
            _fmt(ws, c, _MONEY)
        _fmt(ws, 7, _WHEN)
        _fmt(ws, 11, _PCT)
        _fmt(ws, 22, _WHEN)
        ws.auto_filter.ref = ws.dimensions
        _fit(ws, widths)

    # A workbook that says nothing about where it came from gets emailed on and
    # argued about a week later.
    ws = wb["What happened"]
    ws.insert_rows(1)
    ws["A1"] = (f"{(b.filename if b else 'load')} — load #{batch_id}"
                f"{f', uploaded {b.created_at:%-d %B %Y}' if b and b.created_at else ''}"
                f" — exported {datetime.now():%-d %B %Y, %H:%M}")
    ws["A1"].font = Font(bold=True, size=11)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
