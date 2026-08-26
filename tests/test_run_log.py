"""A record of what the run decided, and why.

    "do we setup a backend log that records all the events when the data is
     loaded and what happens and and pricing and why into an exel"
    "or do you build an api for you too see everything"

Both, off one table. The stages write decisions down as they make them; the
JSON endpoint is how a machine reads them and the workbook is how a person
does. The failure this replaces is specific: every one of these decisions was
already being made and then discarded, so afterwards the only way to answer
"why did it do that" was to re-read the code and guess which branch a row took.

Two properties matter more than the format.

  IT NEVER BREAKS THE RUN. A log that can fail the stage it is logging is worse
  than no log. record() swallows everything.

  IT IS IN WORDS. "3,100 rows had no council valuation, which every valuation
  method needs" is an answer. "rejected: no_cv" is a lookup table somebody has
  to already know, and nobody does at 11pm.
"""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from app.models import (BatchType, ImportBatch, PropertyForSale, RunEvent, User,
                        UserRole, UserStatus)
from app.runlog import events, record


def _batch(db, **kw):
    d = dict(batch_type=BatchType.FOR_SALE.value, region="Auckland",
             filename="auckland.csv", is_active=True, status="published",
             rows_total=100, rows_inserted=60, rows_rejected=40)
    d.update(kw)
    b = ImportBatch(**d)
    db.add(b)
    db.commit()
    return b


@pytest.fixture()
def admin_client(db_session):
    from fastapi.testclient import TestClient

    from app import main
    from app.db import get_db
    from app.security import current_user, require_active, require_admin

    admin = User(email="runlog-admin@test.local", password_hash="x",
                 role=UserRole.ADMIN.value, status=UserStatus.APPROVED.value)
    db_session.add(admin)
    db_session.commit()
    main.app.dependency_overrides = {
        get_db: lambda: db_session, current_user: lambda: admin,
        require_active: lambda: admin, require_admin: lambda: admin,
    }
    try:
        yield TestClient(main.app, raise_server_exceptions=False)
    finally:
        main.app.dependency_overrides = {}


@pytest.fixture()
def anon_client(db_session):
    """No admin override — this is how a stranger reaches the endpoint."""
    from fastapi.testclient import TestClient

    from app import main
    from app.db import get_db

    main.app.dependency_overrides = {get_db: lambda: db_session}
    try:
        yield TestClient(main.app, raise_server_exceptions=False)
    finally:
        main.app.dependency_overrides = {}


def _listing(db, batch, **kw):
    d = dict(import_batch_id=batch.id, address="12 Elliot Street", suburb="Remuera",
             property_type="House", asking_price=1_000_000.0, cv_numeric=1_050_000.0,
             fair_value=1_200_000.0, margin=0.20, confidence="high", comps_used=9,
             floor_area_m2=150.0, is_underpriced=True, is_held=False)
    d.update(kw)
    p = PropertyForSale(**d)
    db.add(p)
    db.commit()
    return p


# ---- the log itself --------------------------------------------------------
def test_a_decision_is_recorded_in_words(db_session):
    b = _batch(db_session)
    record(db_session, stage="load", event="rows_rejected", batch_id=b.id, count=3100,
           detail="3,100 rows rejected — no council valuation, which every "
                  "valuation method needs")

    (e,) = events(db_session, b.id)
    assert e.count == 3100
    assert "council valuation" in e.detail, "the reason must survive as a sentence"


def test_the_log_reads_in_the_order_things_happened(db_session):
    """The order IS the explanation — a rejection at load is why a suburb is
    thin at pricing. Sorted any other way it is a pile of facts."""
    b = _batch(db_session)
    for stage in ("load", "enrich", "price", "publish"):
        record(db_session, stage=stage, event=f"{stage}_done", batch_id=b.id)

    assert [e.stage for e in events(db_session, b.id)] == \
        ["load", "enrich", "price", "publish"]


def test_one_load_never_shows_another_loads_events(db_session):
    a, b = _batch(db_session), _batch(db_session, filename="second.csv")
    record(db_session, stage="load", event="rows_loaded", batch_id=a.id)
    record(db_session, stage="load", event="rows_loaded", batch_id=b.id)

    assert len(events(db_session, a.id)) == 1


def test_the_log_can_never_break_the_run_it_is_logging(db_session):
    """The whole point. A stage that dies because its logging died has been made
    worse by being observed."""
    record(db_session, stage="load", event="x", batch_id=999_999_999)  # no such batch
    record(db_session, stage="load", event="y" * 500, detail="z" * 99_000)
    # Reached here without raising, which is the assertion.


def test_a_long_sentence_is_kept_not_dropped(db_session):
    b = _batch(db_session)
    record(db_session, stage="price", event="priced", batch_id=b.id, detail="w" * 5000)

    (e,) = events(db_session, b.id)
    assert e.detail and len(e.detail) > 1000, "a long reason was thrown away entirely"


def test_warnings_are_marked_so_they_can_be_found(db_session):
    """A run log where nothing stands out is read once and never again."""
    b = _batch(db_session)
    record(db_session, stage="enrich", event="lookups_unreachable", batch_id=b.id,
           level="warn", detail="1,982 lookups never reached the provider")

    assert [e.level for e in events(db_session, b.id)] == ["warn"]


# ---- the API ---------------------------------------------------------------
def test_the_api_returns_the_log_for_the_live_load(admin_client, db_session):
    b = _batch(db_session)
    record(db_session, stage="load", event="rows_loaded", batch_id=b.id, count=60,
           detail="60 rows loaded from auckland.csv, 40 rejected")

    r = admin_client.get("/api/admin/release/run-log")
    assert r.status_code == 200
    assert any("40 rejected" in (e["detail"] or "") for e in r.json())


def test_the_log_is_not_public(anon_client, db_session):
    _batch(db_session)
    assert anon_client.get("/api/admin/release/run-log").status_code in (401, 403)


def test_the_workbook_is_not_public(anon_client, db_session):
    _batch(db_session)
    assert anon_client.get("/api/admin/release/run-log.xlsx").status_code in (401, 403)


# ---- the workbook ----------------------------------------------------------
def _col(ws, name: str) -> int:
    """Find a column by its HEADING. Hard-coded indexes turn every added column
    into a false test failure, which teaches people to edit the numbers without
    reading what broke."""
    for i, c in enumerate(ws[1], start=1):
        if c.value == name:
            return i
    raise AssertionError(f"no {name!r} column — headings are {[c.value for c in ws[1]]}")


def _book(client):
    r = client.get("/api/admin/release/run-log.xlsx")
    assert r.status_code == 200, r.text[:400]
    assert "spreadsheetml" in r.headers["content-type"]
    return load_workbook(io.BytesIO(r.content))


def test_the_workbook_separates_the_five_questions(admin_client, db_session):
    """One sheet answering five questions answers none of them."""
    b = _batch(db_session)
    _listing(db_session, b)
    record(db_session, stage="load", event="rows_loaded", batch_id=b.id, count=60)

    assert _book(admin_client).sheetnames == [
        "What happened", "Why rows dropped", "Every listing", "The deals", "Held back"]


def test_the_deals_sheet_holds_the_deals_and_the_held_sheet_the_held(
        admin_client, db_session):
    b = _batch(db_session)
    _listing(db_session, b)                                     # a deal
    _listing(db_session, b, address="9 Held Road", is_underpriced=False,
             is_held=True, hold_reason="Missing floor area", floor_area_m2=None)

    wb = _book(admin_client)
    assert wb["Every listing"].max_row == 3                     # header + two
    assert wb["The deals"].max_row == 2
    assert wb["Held back"].max_row == 2
    held = wb["Held back"]
    assert held.cell(row=2, column=_col(held, "Why held")).value == "Missing floor area"


def test_the_workbook_says_why_a_listing_has_no_deal(admin_client, db_session):
    """The column this whole change exists for. 317 listings on the last batch
    had a blank margin and nothing anywhere said which rule blanked it."""
    b = _batch(db_session)
    _listing(db_session, b, is_underpriced=False, margin=None,
             deal_block_reason="No advertised price (auction) — the figure in the "
                               "feed is a search price, not what the vendor is asking")

    ws = _book(admin_client)["Every listing"]
    assert "search price" in (ws.cell(row=2, column=_col(ws, "Why no deal")).value or "")


def test_money_reads_as_money_and_dates_as_dates(admin_client, db_session):
    """A spreadsheet that cannot sort its own date column is a screenshot with
    extra steps, and $1200000 is not a number anybody checks at a glance."""
    b = _batch(db_session)
    _listing(db_session, b)

    ws = _book(admin_client)["Every listing"]
    assert ws.cell(row=2, column=_col(ws, "Asking")).number_format == "#,##0"
    assert ws.cell(row=2, column=_col(ws, "Margin %")).number_format == "0.0%"
    when = ws.cell(row=2, column=_col(ws, "Looked up on")).number_format
    assert "yyyy" in when
    assert "yy-mm-dd" not in when, "dates are day, month, year — never the ISO string"


def test_the_funnel_sheet_shows_every_step_not_just_the_total(
        admin_client, db_session):
    b = _batch(db_session)
    _listing(db_session, b)

    ws = _book(admin_client)["Why rows dropped"]
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Listings in this load" in labels
    assert any("flagged as a deal" in (x or "") for x in labels)


def test_every_sheet_of_listings_can_be_filtered(admin_client, db_session):
    """Twenty-one hundred rows without an auto-filter is a wall, not a tool."""
    b = _batch(db_session)
    _listing(db_session, b)

    wb = _book(admin_client)
    for name in ("Every listing", "The deals", "Held back"):
        assert wb[name].auto_filter.ref, f"{name} cannot be filtered"


def test_the_workbook_names_the_load_it_came_from(admin_client, db_session):
    """It gets emailed on and argued about a week later."""
    b = _batch(db_session)
    _listing(db_session, b)

    assert "auckland.csv" in (_book(admin_client)["What happened"]["A1"].value or "")


def test_exporting_with_nothing_loaded_says_so(admin_client):
    r = admin_client.get("/api/admin/release/run-log.xlsx")
    assert r.status_code == 404
    assert "load" in r.json()["detail"].lower()
